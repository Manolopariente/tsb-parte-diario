"""
TSB · Generador de Partes — Motor Final
Estrategia: openpyxl con keep_links=False sobre el template original.
Copia estilos celda a celda con _style para preservar colores indexados.
"""
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, time
import shutil, os

TEMPLATES = {
    "PHZ":  "/home/claude/tsb_generador/TEMPLATE_PHZ_LIMPIO.xlsx",
    "SP":   "/home/claude/tsb_generador/TEMPLATE_SP_LIMPIO.xlsx",
    "CHSN": "/home/claude/tsb_generador/TEMPLATE_CHSN_LIMPIO.xlsx",
}

NOMBRE_YAC = {"PHZ":"PUESTO HERNANDEZ","SP":"SEÑAL PICADA","CHSN":"LOMITA"}
CONTRATO   = "4900100500"

HORARIOS = {
    "PHZ":  {"ini":time(8,0),  "alm_ini":time(13,0),"alm_fin":time(14,0),"fin":time(17,0),  "oc_fin":time(17,15)},
    "SP":   {"ini":time(8,30), "alm_ini":time(13,0),"alm_fin":time(14,0),"fin":time(17,30), "oc_fin":time(17,45)},
    "CHSN": {"ini":time(8,30), "alm_ini":time(13,0),"alm_fin":time(14,0),"fin":time(17,30), "oc_fin":time(17,45)},
}

POSICIONES = {
    1:("H","TARIFA A MT TAREA GENERALES","RH TAR.A P/H CUAD. TAR.GENERALES"),
    3:("H","TARIFA C MT TAREA GENERALES","RH TAR.C P/H CUAD. TAR.GENERALES"),
    9:("H","TARIFA A TOPADORA","SERV. HN TOPAD. 180HP"),
    11:("H","TARIFA C TOPADORA","SERV. HN TOPAD. 180HP"),
    19:("H","TARIFA C MOTONIVELADORA","SERV. TAR.C P/H MOTONIV. 170HP"),
    27:("H","TARIFA C CARGADORA FRONTAL","SERV. TAR.C P/H CARG.FRON. 10TN"),
    35:("H","TARIFA C RETROEXCAVADORA","SERV. TAR.C P/H RETRO 11TN"),
    51:("H","TARIFA C VIBROCOMPACTADOR","SERV. HN V.COMPAC.AUTOIM. 10TN"),
    59:("H","TARIFA C CAMION REGADOR 25M3","SERV. TAR.C P/H CAM.REGAD. 25M3"),
    67:("H","TARIFA C CAMION TRAKKER 18M3","SERV. HN CAM.TRAKKER 18M3"),
    75:("H","TARIFA C CAMION CON BATEA 25M3","SERV. TAR.C P/H CAM.C/BATEA 25M3"),
    81:("DIA","TRACTOR CON CARRETON","SERV. P/DIA TRAC. C/CARRETON"),
    97:("M2","MEMBRANA PEAD 1.0MM","IMPER. P/M2 PILETA MEMB.PEAD 1,0MM"),
    105:("M3","HORMIGON SIMPLE","PROV. IN SITU HORM. H-20"),
    107:("KG","PROVISION DE ARMADURA","PROV. IN SITU ARM.HIERRO P/H°A°"),
    117:("M2","PROV Y COLOC MALLA ANTIPAJARO","PROV/INST IN SITU MALLA ANTIPAJARO"),
    181:("H","TARIFA A CAMION CON BATEA 25M3","SERV. TAR.A P/H CAM.C/BATEA 25M3"),
    197:("H","TARIFA C EXCAVADORA","SERV. TAR.C EXCAVAD. 20TN"),
    207:("H","TARIFA A CAMION REGADOR 25M3","SERV. TAR.A P/H CAM.REGAD. 25M3"),
    215:("H","TARIFA A CAMION TRAKKER 18M3","SERV. HN CAM.TRAKKER 18M3"),
    225:("H","TARIFA A CARGADORA FRONTAL","SERV. TAR.A P/H CARG.FRON. 10TN"),
    233:("H","TARIFA A MOTONIVELADORA","SERV. TAR.A P/H MOTONIV. 170HP"),
    237:("H","TARIFA A RETROEXCAVADORA","SERV. TAR.A P/H RETRO 11TN"),
    243:("H","TARIFA A MT TAREA GENERALES","RH TAR.A P/H CUAD. TAR.GENERALES"),
    257:("H","TARIFA A EXCAVADORA","SERV. TAR.A EXCAVAD. 20TN"),
}

EQUIPOS = {
    "PHZ":[
        {"id_ypf":"4659 -Atg",       "interno":226,  "pos_a":243,"pos_c":3,   "tg":True},
        {"id_ypf":"4656 -Atg",       "interno":5154, "pos_a":243,"pos_c":3,   "tg":True},
        {"id_ypf":"4663 -Retro",     "interno":7609, "pos_a":237,"pos_c":35},
        {"id_ypf":"4664 -Retro",     "interno":7613, "pos_a":237,"pos_c":35},
        {"id_ypf":"4689 -Cargadora", "interno":7604, "pos_a":225,"pos_c":27},
        {"id_ypf":"4669 -Moto",      "interno":7616, "pos_a":233,"pos_c":19},
        {"id_ypf":"4687 -traker",    "interno":9226, "pos_a":215,"pos_c":67},
        {"id_ypf":"4693 -Topadora",  "interno":5736, "pos_a":None,"pos_c":11, "sinprog":True},
        {"id_ypf":"4680 -Batea",     "interno":1472, "pos_a":None,"pos_c":75, "sinprog":True},
        {"id_ypf":"4691 -Vibro",     "interno":7630, "pos_a":None,"pos_c":51, "sinprog":True},
        {"id_ypf":"4692 -Vibro",     "interno":7631, "pos_a":None,"pos_c":51, "sinprog":True},
    ],
    "SP":[
        {"id_ypf":"5555 -Atg",            "interno":197,  "pos_a":243,"pos_c":3,   "tg":True},
        {"id_ypf":"4666 -retro",          "interno":7614, "pos_a":237,"pos_c":35},
        {"id_ypf":"4671 -motoniveladora", "interno":7563, "pos_a":233,"pos_c":19},
        {"id_ypf":"4683 -batea",          "interno":1391, "pos_a":181,"pos_c":75},
        {"id_ypf":"4678 -excavadora",     "interno":9704, "pos_a":257,"pos_c":197,"condicional":True},
    ],
    "CHSN":[
        {"id_ypf":"4657 -ATG",           "interno":174,  "pos_a":243,"pos_c":3,   "tg":True},
        {"id_ypf":"4673 -Regador 35m3",  "interno":1552, "pos_a":207,"pos_c":59},
        {"id_ypf":"4674 -Regador 35m3",  "interno":1566, "pos_a":207,"pos_c":59},
        {"id_ypf":"4668 -Motoniveladora","interno":7564, "pos_a":233,"pos_c":19},
        {"id_ypf":"4690 -Cargadora",     "interno":7576, "pos_a":225,"pos_c":27},
        {"id_ypf":"4679 -excavadora",    "interno":7596, "pos_a":257,"pos_c":197},
        {"id_ypf":"5255 -Motoniveladora","interno":7618, "pos_a":233,"pos_c":19},
        {"id_ypf":"4662 -Retro",         "interno":7624, "pos_a":237,"pos_c":35},
        {"id_ypf":"4684 -Tracker",       "interno":9228, "pos_a":215,"pos_c":67},
        {"id_ypf":"4685 -Tracker",       "interno":9238, "pos_a":215,"pos_c":67},
    ],
}

ONCALL_ITEMS = {
    "hormigon":(105,"M3"), "malla":(117,"M2"),
    "geomembrana":(97,"M2"), "hierro":(107,"KG"),
}

CONDUCTORES = [
    "DIAZ CARLOS RAUL","POULAGE JULIO HECTOR","VALENZUELA CARLOS RAUL","VILLAROEL DANIEL ANGEL",
    "COFRE JAVIER NICOLAS","GUAJARDO SERGIO DANIEL","VIZCARRA WALTER ANDRES","ZOZAYA HECTOR JULIAN",
    "NAHUELÑIR JULIO ALBERTO","ZUÑIGA JORGE DANIEL","ZUÑIGA CLAUDIO ALEJANDRO","HERNANDEZ HECTOR ADRIAN",
    "PACHECO JORGE DANIEL","AVILA DANIEL ANTONIO","MUÑOZ DAVID","TORRES DARDO EZEQUIEL",
    "PROCOPIO DE ALMEIDA ANTONIO EDUARDO","CHIRINO GABRIEL","HERNANDEZ FRANCISCO JAIME",
    "SUAREZ FABIAN ERNESTO","POGGI MAURICIO IVAN","BRAVO MATIAS ALEJANDRO","GUAJARDO ALEJANDRO ANTONIO",
    "VILLALVA JAVIER FRANCISCO","MOLINA PEDRO SEBASTIAN","RIQUELME FRANCO EMANUEL","VALDEZ DIEGO ALBERTO",
    "ETAYO ALEJANDRO GONZALO","KEODUANGSA PAIWAN","ALFARO CLAUDIO JAVIER","GALVAN JOSE FERNANDO",
    "PARIENTE MANUEL ANGEL","ARANGO EMMA MARISA","CANTERO MARCOS JAVIER","GOMEZ PABLO JAVIER",
    "GORDILLO LUIS ALBERTO","VILLALBA ANIBAL","PEREZ MARCOS EXEQUIEL","POBLETE CESAR",
    "PERALTA JUAN CARLOS","MARANGEL SANTANA CARLOS OMAR","FUSTER DIEGO HERNAN","JARA GUSTAVO ADRIANO",
    "GUERRA OSCAR MARCELO","QUIROGA JUAN JOSE","ESPINOZA JUAN RAMON","PEÑALVE IGNACIO BENJAMIN",
    "HERNANDEZ JOAN MANUEL","LAGOS FACUNDO LUIS ALBERTO","JARA PLACIDO MARIO ALBERTO","PINEDA CESAR EDUARDO",
    "PADILLA AGUSTIN LEONARDO","CERDA NESTOR SEGUNDO","REVECO LUIS EDUARDO","VELAZQUEZ ELIAS MATEO",
    "LOPEZ GABRIEL OSCAR","URPIANELLO RODRIGO ALEJANDRO","PENAYO GUSTAVO HORACIO","LLANTEN GERARDO MARTIN MANUEL",
    "MESA GONZALO NICOLAS","YOBANOLO VANESA LILIANA",
]

def hs(t): return t.hour*3600/86400 + t.minute*60/86400

def get_pos(eq, act):
    if eq.get("sinprog"): return eq["pos_c"]
    if act == "Tarea": return eq.get("pos_a")
    if act == "Falla Equipo": return None
    return eq["pos_c"]

def pos_info(pos):
    return POSICIONES.get(pos, ("","",""))

def copiar_estilo(src, dst):
    """Copia el _style completo — preserva colores indexados."""
    dst._style = src._style
    dst.number_format = src.number_format

def escribir_fila(ws, row, datos, fila_base):
    """Escribe datos copiando estilos de fila_base."""
    ws.row_dimensions[row].height = 30
    for col in range(1, 27):
        dst = ws.cell(row=row, column=col)
        src = ws.cell(row=fila_base, column=col)
        copiar_estilo(src, dst)
        if col in datos and datos[col] is not None:
            dst.value = datos[col]
        else:
            dst.value = None

def t2excel(t): return t.hour/24 + t.minute/1440

def generar_parte(yacimiento, fecha_parte, supervisor,
                  equipos_dia, conductores_dia,
                  actividades_dia=None, ots_dia=None,
                  oncall=None, tg_oncall=None):

    if actividades_dia is None: actividades_dia = {}
    if ots_dia is None: ots_dia = {}
    if oncall is None: oncall = {}

    # Cargar template original con keep_links=False
    wb = load_workbook(TEMPLATES[yacimiento], keep_links=False)
    ws = wb['Parte Diario']

    # Limpiar filas 8+ — guardar estilos de filas base primero
    # (no eliminar, solo limpiar valores)
    max_row = ws.max_row
    for r in range(8, max_row+1):
        for c in range(1, 27):
            ws.cell(row=r, column=c).value = None

    # Actualizar cabecera
    ws['C3'] = fecha_parte.strftime('%d/%m/%Y')
    ws['G4'] = supervisor

    hor = HORARIOS[yacimiento]
    row = 8

    def make_datos(interno, id_ypf, act, ot_info, conductor, ini, fin, pos, falla=False, es_alm=False, f_cant=None):
        um, desc_pos, _ = pos_info(pos) if pos else ("","","")
        ini_dec = t2excel(ini)
        fin_dec = t2excel(fin)
        hs_val  = round((fin_dec - ini_dec) * 24, 4)
        return {
            1:  interno,
            2:  NOMBRE_YAC[yacimiento],
            3:  "MOV._DE_SUELO",
            4:  "OTROS MOV. SUELOS",
            5:  um if not falla else "",                  # E = UM (valor)
            6:  f_cant,                                    # F = cantidad oncall
            7:  id_ypf,                                    # G = ID YPF (valor)
            9:  round(hs_val,4) if act=="Tarea" and not falla and not es_alm else 0,  # I = TNR
            10: round(hs_val,4) if act not in ["Tarea","Almuerzo"] and not falla else 0,  # J = TTA
            11: act,                                        # K = Actividad
            12: "refrigerio" if es_alm else ot_info.get("descripcion",""),  # L
            13: ot_info.get("ot","") if not falla else "", # M = OT
            14: ot_info.get("operacion","") if not falla else "",  # N = Operacion
            15: conductor,                                  # O = Conductor
            16: ini_dec,                                    # P = Desde
            17: fin_dec,                                    # Q = Hasta
            18: ot_info.get("ubicacion","") if not falla else "",  # R = Ubicacion
            19: pos if not falla else None,                # S = Posicion
            20: 10 if not falla else None,                 # T = Linea
            21: f_cant if f_cant else (0 if es_alm or falla else round(hs_val,4)),  # U = Cantidad
            22: um if not falla else "",                   # V = UM
            23: desc_pos if not falla else "",             # W = Desc posicion
            24: 8 if act=="Tarea" and not falla else 0,   # X = KPI
            25: fecha_parte.strftime('%d/%m/%Y'),           # Y = Fecha
            26: CONTRATO,                                   # Z = Contrato
        }

    equipos_activos = [e for e in EQUIPOS[yacimiento] if e["id_ypf"] in equipos_dia and not e.get("sinprog")]
    equipos_sinprog = [e for e in EQUIPOS[yacimiento] if e.get("sinprog")]

    for eq in equipos_activos:
        id_ypf   = eq["id_ypf"]
        interno  = eq["interno"]
        es_tg    = eq.get("tg", False)
        conductor = conductores_dia.get(id_ypf, "")
        act_info  = actividades_dia.get(id_ypf, {})
        act_man   = act_info.get("mañana","Tarea")
        act_tar   = act_info.get("tarde","Tarea")

        ots_eq = ots_dia.get(id_ypf, [{"ot":"","operacion":"","ubicacion":""}])
        n = len(ots_eq)

        if n == 1:
            bloques_man = [{"ini":hor["ini"],"fin":hor["alm_ini"],"ot":ots_eq[0],"act":act_man}]
            bloques_tar = [{"ini":hor["alm_fin"],"fin":hor["fin"],"ot":ots_eq[0],"act":act_tar}]
        elif n == 2:
            bloques_man = [{"ini":hor["ini"],"fin":hor["alm_ini"],"ot":ots_eq[0],"act":act_man}]
            bloques_tar = [{"ini":hor["alm_fin"],"fin":hor["fin"],"ot":ots_eq[1],"act":act_tar}]
        elif n == 3:
            bloques_man = [
                {"ini":hor["ini"],"fin":time(10,0),"ot":ots_eq[0],"act":act_man},
                {"ini":time(10,0),"fin":hor["alm_ini"],"ot":ots_eq[1],"act":act_man},
            ]
            bloques_tar = [{"ini":hor["alm_fin"],"fin":hor["fin"],"ot":ots_eq[2],"act":act_tar}]
        else:
            bloques_man = [
                {"ini":hor["ini"],"fin":time(9,0),"ot":ots_eq[0],"act":act_man},
                {"ini":time(9,0),"fin":time(11,0),"ot":ots_eq[1],"act":act_man},
                {"ini":time(11,0),"fin":hor["alm_ini"],"ot":ots_eq[2],"act":act_man},
            ]
            bloques_tar = [{"ini":hor["alm_fin"],"fin":hor["fin"],"ot":ots_eq[3],"act":act_tar}]

        for b in bloques_man:
            pos = get_pos(eq, b["act"])
            falla = b["act"] == "Falla Equipo"
            datos = make_datos(interno,id_ypf,b["act"],b["ot"],conductor,b["ini"],b["fin"],pos,falla)
            escribir_fila(ws, row, datos, 8)
            row += 1

        # Almuerzo
        datos_alm = make_datos(interno,id_ypf,"Almuerzo",{},conductor,hor["alm_ini"],hor["alm_fin"],None,es_alm=True)
        escribir_fila(ws, row, datos_alm, 9)
        row += 1

        for b in bloques_tar:
            pos = get_pos(eq, b["act"])
            falla = b["act"] == "Falla Equipo"
            datos = make_datos(interno,id_ypf,b["act"],b["ot"],conductor,b["ini"],b["fin"],pos,falla)
            escribir_fila(ws, row, datos, 10)
            row += 1

        # On-call pegado a TG
        if es_tg and (not tg_oncall or tg_oncall == id_ypf):
            for key,(pos_oc,um_oc) in ONCALL_ITEMS.items():
                cant = oncall.get(key)
                if cant:
                    d_oc = make_datos(interno,id_ypf,"Tarea",{},conductor,hor["fin"],hor["oc_fin"],pos_oc,f_cant=cant)
                    escribir_fila(ws, row, d_oc, 10)
                    row += 1

        # 2 filas vacías con estilo
        for _ in range(2):
            ws.row_dimensions[row].height = 30
            for c in range(1,27):
                copiar_estilo(ws.cell(row=8,column=c), ws.cell(row=row,column=c))
                ws.cell(row=row,column=c).value = None
            row += 1

    # Carretón
    if oncall.get("carreton"):
        um_c,desc_c,_ = pos_info(81)
        d_car = make_datos(1709,"4694 -Carretón","Tarea",{},None,hor["ini"],hor["fin"],81,f_cant=1)
        escribir_fila(ws, row, d_car, 8)
        row += 1
        for _ in range(2):
            ws.row_dimensions[row].height = 30
            for c in range(1,27):
                copiar_estilo(ws.cell(row=8,column=c), ws.cell(row=row,column=c))
                ws.cell(row=row,column=c).value = None
            row += 1

    # Sin programación
    for eq in equipos_sinprog:
        for fb, act, ini, fin in [
            (8,"Espera OT",hor["ini"],hor["alm_ini"]),
            (9,"Almuerzo",hor["alm_ini"],hor["alm_fin"]),
            (10,"Espera OT",hor["alm_fin"],hor["fin"]),
        ]:
            es_alm = act == "Almuerzo"
            pos = eq["pos_c"] if not es_alm else None
            datos = make_datos(eq["interno"],eq["id_ypf"],act,{},None,ini,fin,pos,es_alm=es_alm)
            if not es_alm: datos[12] = "Sin programacion"
            escribir_fila(ws, row, datos, fb)
            row += 1
        for _ in range(2):
            ws.row_dimensions[row].height = 30
            for c in range(1,27):
                copiar_estilo(ws.cell(row=8,column=c), ws.cell(row=row,column=c))
                ws.cell(row=row,column=c).value = None
            row += 1

    # Q5 y Print_Area dinámicos
    ultima = row - 1
    ws['Q5'] = f"=SUM(U8:U{ultima})"

    # Eliminar hojas auxiliares
    for shname in list(wb.sheetnames):
        if shname not in ['Parte Diario']:
            del wb[shname]

    # Agregar hoja Datos
    wd = wb.create_sheet("Datos")
    _construir_datos(wd)

    # Ajustar Print_Area
    wb.defined_names['_xlnm.Print_Area'] = openpyxl.workbook.defined_name.DefinedName(
        '_xlnm.Print_Area',
        attr_text=f"'Parte Diario'!$A$1:$Z${ultima}"
    )

    return wb


def _construir_datos(wd):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    FN = Font(name="Arial",size=9)
    FB = Font(name="Arial",size=9,bold=True,color="FFFFFF")
    FH = PatternFill("solid",start_color="1F3864")
    FS = PatternFill("solid",start_color="2E75B6")
    BD = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))

    def sec(r,t):
        c=wd.cell(row=r,column=1); c.value=t; c.font=FB; c.fill=FH
    def hdr(r,vals):
        for i,v in enumerate(vals):
            c=wd.cell(row=r,column=i+1); c.value=v; c.font=FB; c.fill=FS; c.border=BD
    def fila(r,vals):
        for i,v in enumerate(vals):
            c=wd.cell(row=r,column=i+1); c.value=v; c.font=FN; c.border=BD

    sec(1,"MAESTRO DE EQUIPOS")
    hdr(2,["ID YPF","INTERNO TSB","YACIMIENTO","POS TAR A","POS TAR C"])
    r=3
    for yac,eqs in EQUIPOS.items():
        for eq in eqs:
            fila(r,[eq["id_ypf"],eq["interno"],NOMBRE_YAC[yac],eq.get("pos_a",""),eq.get("pos_c","")]); r+=1
    r+=2; sec(r,"TABLA DE POSICIONES")
    r+=1; hdr(r,["POSICION","LINEA","UM","DESCRIPCION","TEXTO BREVE 2"])
    r+=1
    for pos,(um,t1,t2) in sorted(POSICIONES.items()):
        fila(r,[pos,10,um,t1,t2]); r+=1
    r+=2; sec(r,"CONDUCTORES")
    r+=1
    for nombre in CONDUCTORES:
        c=wd.cell(row=r,column=1); c.value=nombre; c.font=FN; c.border=BD; r+=1

    for cl,w in [('A',25),('B',12),('C',20),('D',10),('E',10)]:
        wd.column_dimensions[cl].width=w


# TEST
if __name__ == "__main__":
    wb = generar_parte(
        yacimiento="PHZ", fecha_parte=date(2026,5,17),
        supervisor="PARIENTE MANUEL ANGEL",
        equipos_dia=["4659 -Atg","4656 -Atg","4663 -Retro","4664 -Retro","4689 -Cargadora","4669 -Moto","4687 -traker"],
        conductores_dia={},
        ots_dia={
            "4659 -Atg":      [{"ot":"2008880559","operacion":"10", "ubicacion":"CV-NEN-Y-PHM-BATPHZ09-COLENT01-PH3006"}],
            "4656 -Atg":      [{"ot":"2008816913","operacion":"330","ubicacion":"CV-NEN-P-PHM-PTAPHZ04"}],
            "4663 -Retro":    [{"ot":"2008816913","operacion":"320","ubicacion":"CV-NEN-P-PHM-PTAPHZ04"}],
            "4664 -Retro":    [{"ot":"2008880559","operacion":"5",  "ubicacion":"CV-NEN-Y-PHM-BATPHZ09-COLENT01-PH3006"}],
            "4689 -Cargadora":[{"ot":"2008793997","operacion":"280","ubicacion":"CV-NEN-P-PHM-PIAPHZ02-SATAUX29-PH0324"}],
            "4669 -Moto":     [{"ot":"2008793997","operacion":"250","ubicacion":"CV-NEN-P-PHM-PIAPHZ02-SATAUX29-PH0324"}],
            "4687 -traker":   [{"ot":"2008793997","operacion":"270","ubicacion":"CV-NEN-P-PHM-PIAPHZ02-SATAUX29-PH0324"}],
        },
        oncall={}
    )
    out = "/home/claude/tsb_generador/4900100500_17052026_PD_PHZ_FINAL.xlsx"
    wb.save(out)
    print(f"✅ {out}")
    print(f"Hojas: {wb.sheetnames}")


# ============================================================
# EXTRACTOR DE OTs
# ============================================================
def extraer_info_ot(pdf_path):
    """Extrae OT, equipo, operación, descripción y ubicación de un PDF de OT."""
    import fitz, re
    doc = fitz.open(pdf_path)
    text = doc[0].get_text()
    doc.close()
    
    ot_match = re.search(r'Orden de Trabajo N[°o]\s+(\d+)', text)
    ot_num = ot_match.group(1) if ot_match else ""
    
    eq_match = re.search(r'(\d{4})\s*-\s*(?:ATT|MOT|REX|CFR|CTR|REG)[_A-Z0-9]+', text)
    eq_id = eq_match.group(1) if eq_match else ""
    
    matches = re.findall(r'Operaci[oó]n\s+(\d+)\s+-\s+(.+?)\s+Programaci[oó]n', text, re.DOTALL)
    if matches:
        op_num, op_desc = matches[-1]
        op_desc = re.sub(r'\s+', ' ', op_desc).strip()
    else:
        op_num, op_desc = "", ""
    
    ubi_match = re.search(r'C[oó]digo\s+(CV-[A-Z0-9\-]+)', text)
    ubicacion = ubi_match.group(1) if ubi_match else ""
    
    return {
        "ot": ot_num,
        "eq_id": eq_id,
        "operacion": op_num,
        "descripcion": op_desc,
        "ubicacion": ubicacion,
    }


def procesar_ots_pdf(pdf_paths, equipos_yac):
    """
    Recibe lista de PDFs de OTs y devuelve dict listo para ots_dia.
    Detecta automáticamente a qué equipo pertenece cada OT.
    Respeta el orden de los equipos en el parte.
    """
    # Extraer info de cada PDF
    ots_extraidas = []
    for path in pdf_paths:
        info = extraer_info_ot(path)
        ots_extraidas.append(info)
    
    # Mapear eq_id (ej: "4669") → id_ypf (ej: "4669 -Moto")
    id_map = {}
    for eq in equipos_yac:
        # Extraer número del id_ypf (ej: "4669 -Moto" → "4669")
        import re
        m = re.match(r'(\d+)', eq["id_ypf"])
        if m:
            id_map[m.group(1)] = eq["id_ypf"]
    
    # Construir ots_dia
    ots_dia = {}
    for ot in ots_extraidas:
        id_ypf = id_map.get(ot["eq_id"])
        if id_ypf:
            if id_ypf not in ots_dia:
                ots_dia[id_ypf] = []
            ots_dia[id_ypf].append({
                "ot":          ot["ot"],
                "operacion":   ot["operacion"],
                "ubicacion":   ot["ubicacion"],
                "descripcion": ot["descripcion"],
            })
    
    return ots_dia
